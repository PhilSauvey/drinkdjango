import openpyxl
import google.appengine

print "done"

"""

from openpyxl import load_workbook
book = load_workbook('database.xlsx')
ingsheet = book['Sheet2']
typesheet=book['Sheet3']
sheet = book['Sheet1']
for i in range(1,typesheet.max_row+1):
	t=DrinkType()
	t.drink_type=str(typesheet.cell(row=i,column=1).value).encode()
	t.put()
for j in range (2,sheet.max_row+1):
	type_key=ndb.Key(DrinkType,str(sheet.cell(row=j,column=4).value).encode())
	a = Drink(parent=type_key)
	a.drink_name=str(sheet.cell(row=j,column=1).value).encode()
	a.drink_glass=str(sheet.cell(row=j,column=7).value).encode()
	a.drink_garnish=str(sheet.cell(row=j, column=5).value).encode()
	a.drink_inst=str(sheet.cell(row=j, column=6).value).encode()
	a.drink_slug=str(sheet.cell(row=j, column=8).value).encode()
	drink_key=a.put()	
	for i in range(9,sheet.max_column+1):
		if sheet.cell(row=j, column=i).value !=None:
			n=Ingredients(parent=drink_key)
			n.ing_name=str(sheet.cell(row=1, column=i).value).encode()
			n.ing_amount=str(sheet.cell(row=j, column=i).value).encode()
			n.put()
index=0
for k in range(1,ingsheet.max_row+1):
	q= AllIngredients(ing_name=str(ingsheet.cell(row=k, column=1).value).encode())
	q.index=index			
	q.put()
	index+=1"""