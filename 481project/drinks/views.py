from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, Http404
from django.template import loader
from .models import Drink, Ingredients, AllIngredients, User, DrinkLink, UserIng, DrinkType
import operator
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.utils.datastructures import MultiValueDictKeyError
from django.contrib.auth import logout
from django.contrib import auth
from django.http import HttpResponseRedirect
from google.appengine.ext import ndb

def ing_check(data,owned):
	for m in data:
		if m.ing_name not in owned:
			return False
	return True

def login_check(request):
	state=0
	if request.method=="POST":
		try:
			username = request.POST["username"]
		except (MultiValueDictKeyError):
			try:
				log = request.POST["logout"]
			except (MultiValueDictKeyError):
				try:
					add_drink = request.POST["add_drink"]
				except:
					state=0
				else:
					state=4
			else:
				state=3
		else:
			try:
				password = request.POST["password"]
			except (MultiValueDictKeyError):
				state=0
			else:
				state=1
	return state
	
def index(request):
	
	types=DrinkType.query().fetch()
	types.sort(key=lambda x:x.drink_type)
	drink_lists = ()
	try:
		type = request.GET["type"]
	except (MultiValueDictKeyError):
		drink_lists = ()
	else:
		type=DrinkType.query().filter(DrinkType.drink_type==type).fetch()
		if type:
			type=type[0]
			type_key=ndb.Key(DrinkType,str(type))
			list=Drink.query(ancestor=type_key).fetch()
			list.sort(key=lambda x: x.drink_name)
			drink_lists=(type,list)
	
	state=login_check(request)
	context = {"drink_lists":drink_lists,"types":types}
	response=render(request,"drinks/index.html",context)
	
	if 'user' in request.COOKIES:
		user=User.query().filter(User.username==request.COOKIES['user']).fetch()
		if state==3:
			response.delete_cookie('user')
			response.delete_cookie('searched')
			response.delete_cookie('drink')
		elif user:
			user=user[0]
			context["username"]=user
			response=render(request,"drinks/index.html",context)
	else:
		if state==1:
			username = request.POST["username"]
			password = request.POST["password"]
			if User.query().filter(User.username==username).filter(User.password==password).fetch():
				user=User.query().filter(User.username==username).filter(User.password==password).fetch()[0]
				context["username"]=user
				response=render(request,"drinks/index.html",context)
				response.set_cookie('user', username)
			else:
				message="Invalid Username or Password"
				context["message"]=message
				response=render(request,"drinks/index.html",context)
	
	return response
# Create your views here.


def detail(request, drink_slug):
	glasses={
		"Pint":"/static/drinks/Pint_Glass.png",
		"Highball":"/static/drinks/Highball_Glass.png",
		"Sour Glass":"/static/drinks/Sour_Glass.png",
		"Martini":"/static/drinks/Cocktail_Glass.png",
		"Champagne":"/static/drinks/Champagne_Glass.png",
		"Old Fashioned":"/static/drinks/Old_Fashioned_Glass.png",
		"Wine":"/static/drinks/Wine-glass.png",
		"Wineglass":"/static/drinks/Wine-glass.png",
		"Collins":"/static/drinks/Collins_Glass.png",
		"Champagne Coupe":"/static/drinks/Champagne_Coupe_Glass.png",
		"Pousse-Cafe":"/static/drinks/pousse_cafe_glass.png",
		"Cordial":"/static/drinks/Cordial_Glass.png",
		"Double Old Fashioned":"/static/drinks/Old_Fashioned_Glass.png",
		"Hurricane":"/static/drinks/Hurricane_Glass.png",
		"Large Wine":"/static/drinks/Wine-glass.png",
		"Margarita":"/static/drinks/Margarita_Glass.png",
		"Shot":"/static/drinks/shotglass.png",
		"Pitcher":"/static/drinks/pitcher.png",
		"Mule":"/static/drinks/mule_glass.png",
		"Rocks":"/static/drinks/Rocks_Glass.png",
		"Coupe":"/static/drinks/Coupe_Glass.png",
		"Jug":"/static/drinks/jug.png",
		"Tea Cup":"/static/drinks/teacup.png"	
	}
	state = login_check(request)
	drink = Drink.query().filter(Drink.drink_slug==drink_slug).fetch()[0]
	drink_key=drink.key
	ing_list=Ingredients.query(ancestor=drink_key).fetch()
	if drink.drink_glass=="See Note":
		image_url = ["/static/drinks/Rocks_Glass.jpg","/static/drinks/Cocktail_Glass.png"]
	else:
		image_url = [glasses[drink.drink_glass]]
	response = render(request, "drinks/detail.html",{"drink":drink,"ing_list":ing_list,"img_url":image_url})
	context={"drink":drink,"ing_list":ing_list,"img_url":image_url}
	if state==3:
		response.delete_cookie('user')
		response.delete_cookie('searched')
		response.delete_cookie('drink')
	elif 'user' in request.COOKIES:
		user=User.query().filter(User.username==request.COOKIES['user']).fetch()
		if user:
			user=user[0]
			context["username"]=user
		if state==4:
			if not DrinkLink.query(DrinkLink.user_key==user.key,DrinkLink.drink_key==drink_key).fetch():
				link=DrinkLink()
				link.user_key=user.key
				link.drink_key=drink_key
				link.put()
				message="You have added "+drink.drink_name+" to your saved drink list"
			else:
				message="That drink is already on your list"
			context["message"]=message	
		response=render(request, "drinks/detail.html",context)
		response.set_cookie('drink',drink_slug)
		
	else:
		if state==1:
			username = request.POST["username"]
			password = request.POST["password"]
			if User.query().filter(User.username==username).filter(User.password==password).fetch():
				user=User.query().filter(User.username==username).filter(User.password==password).fetch()[0]
				response=render(request, "drinks/detail.html",{"drink":drink,"ing_list":ing_list,"username":user,"img_url":image_url})
				response.set_cookie('user', username)
			else:
				message="Invalid Username or Password"
				context["message"]=message
				response = render(request, "drinks/detail.html",context)
			
	return response

def results(request):
	owned_list=[]
	make_list=[]
	prev_list=[]
	error=1
	user=""
	if 'user' in request.COOKIES:
		user=User.query().filter(User.username==request.COOKIES['user']).fetch()
		if user:
			user=user[0]
			user_ing=UserIng.query(ancestor=user.key).fetch()
			if request.method=="POST":
				for ing in user_ing:
					ing.key.delete()
			else:
				for ing in user_ing:
					prev_list.append(ing.ing_name)
	drink_list = Drink.query().fetch()
	ing_list = AllIngredients.query().fetch()
	for ingredients in ing_list:
		if ingredients.ing_name in request.POST or ingredients.ing_name in prev_list:
			owned_list.append(ingredients.ing_name)
			if user:
				ing=UserIng(parent=user.key)
				ing.ing_name=ingredients.ing_name
				ing.put()
	if len(owned_list)==0:
		ing_list.sort(key=lambda x: x.index)
		return render(request, "drinks/search.html", {
			"ing_list":ing_list,
			"error_message": "You didn't select any ingredients.",})
	else:
		missing_list=[]
		buy_list={}
		for drinks in drink_list:
			unmakeable=0
			missing_ing=[]
			drink_key=drinks.key
			ing_list=Ingredients.query(ancestor=drink_key).fetch()
			for i in ing_list: 
				if i.ing_name not in owned_list:
					unmakeable=1
					missing_ing.append(i.ing_name)
			missing_list.append(missing_ing)
			if unmakeable==0:
				make_list.append(drinks)
		for list in missing_list:
			if len(list)==1:
				if list[0] in buy_list.keys():
					buy_list[list[0]]+=1
				else:
					buy_list[list[0]]=1

	
		sorted_buy = sorted(buy_list.items(), key=operator.itemgetter(1))
		
	buy=sorted_buy[len(sorted_buy)-1]
	make_list.sort(key=lambda x:x.drink_name)
	response=render(request, "drinks/results.html",{"make_list":make_list,"missing_list":missing_list,"buy":buy},)
	if user:
		response.set_cookie('searched',"True")
	return response

def search(request):
	ing_list = AllIngredients.query().fetch()
	ing_list.sort(key=lambda x: x.index)
	user_list=[]
	context = {"ing_list":ing_list}
	if 'user' in request.COOKIES:
		user=User.query().filter(User.username==request.COOKIES['user']).fetch()[0]
		user_ing=UserIng.query(ancestor=user.key).fetch()

		for ing in user_ing:
			user_list.append(ing.ing_name)	
	context["user_ing"]=user_list
	return render(request, "drinks/search.html",context)

def mydrinks(request):
	if 'user' in request.COOKIES:
		user=User.query().filter(User.username==request.COOKIES['user']).fetch()[0]
		drink_list=DrinkLink.query(DrinkLink.user_key==user.key).fetch()
		drinks=[]
		for obj in drink_list:
			drinks.append(obj.drink_key.get())
		context={"username":user,"drinks":drinks, 'drink_list':drink_list}
		return render(request, "drinks/mydrinks.html",context)
	else:
		return redirect("drinks/")

def createUser(request):
	context={}
	status=0
	if request.method=="POST":
		try:
			username=request.POST["username"]
		except(MultiValueDictKeyError):
			context["message"]="Please Enter A Username"
		else:
			if User.query().filter(User.username==username).fetch():
				context["message"]="That Username is Already Taken, Please Enter a New Username"
			else:
				try:
					password=request.POST["password"]
				except:
					context["message"]="Please Enter a Password"
				else:
					u=User()
					u.username=username
					u.password=password
					u.put()
					status=1
					context["username"]=u
					context["message"]="You have successfully created a new user, and are now signed in."
	
	response=render(request, "drinks/newuser.html",context)
	if status==1:
		response.set_cookie('user', username)
	return response
		
def populate(request):
	import openpyxl
	from openpyxl import load_workbook
	book = load_workbook('database.xlsx')
	ingsheet = book['Sheet2']
	typesheet=book['Sheet3']
	sheet = book['Sheet1']
	for number in range(692,792):
		type_key=ndb.Key(DrinkType,str(sheet.cell(row=number,column=4).value).encode())
		a = Drink(parent=type_key)
		a.drink_name=str(sheet.cell(row=number,column=1).value).encode()
		a.drink_glass=str(sheet.cell(row=number,column=7).value).encode()
		a.drink_garnish=str(sheet.cell(row=number, column=5).value).encode()
		a.drink_inst=str(sheet.cell(row=number, column=6).value).encode()
		a.drink_slug=str(sheet.cell(row=number, column=8).value).encode()
		drink_key=a.put()
		for i in range(9,sheet.max_column+1):
			if sheet.cell(row=number, column=i).value !=None:
				n=Ingredients(parent=drink_key)
				n.ing_name=str(sheet.cell(row=1, column=i).value).encode()
				n.ing_amount=str(sheet.cell(row=number, column=i).value).encode()
				n.put()	
	context={}	
	response=render(request, "drinks/populate.html",context)
	return response