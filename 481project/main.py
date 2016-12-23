#!/usr/bin/env python
#
# Copyright 2007 Google Inc.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

from django.core.wsgi import get_wsgi_application
from google.appengine.ext import ndb

import openpyxl
	from openpyxl import load_workbook
	book = load_workbook('database.xlsx')
	ingsheet = book['Sheet2']
	typesheet=book['Sheet3']
	sheet = book['Sheet1']
	for i in range(1,typesheet.max_row+1):
		t=DrinkType()
		t.drink_type=str(typesheet.cell(row=i,column=1).value).encode()
		t.put()
	for j in range (2,sheet.max_row+1):
		type_key=ndb.Key(DrinkType,str(sheet.cell(row=j,column=4).value).encode())
		a = Drink(parent=type_key)
		a.drink_name=str(sheet.cell(row=j,column=1).value).encode()
		a.drink_glass=str(sheet.cell(row=j,column=7).value).encode()
		a.drink_garnish=str(sheet.cell(row=j, column=5).value).encode()
		a.drink_inst=str(sheet.cell(row=j, column=6).value).encode()
		a.drink_slug=str(sheet.cell(row=j, column=8).value).encode()
		drink_key=a.put()	
		for i in range(9,sheet.max_column+1):
			if sheet.cell(row=j, column=i).value !=None:
				n=Ingredients(parent=drink_key)
				n.ing_name=str(sheet.cell(row=1, column=i).value).encode()
				n.ing_amount=str(sheet.cell(row=j, column=i).value).encode()
				n.put()
			
	index=0
	for k in range(1,ingsheet.max_row+1):
		q= AllIngredients(ing_name=str(ingsheet.cell(row=k, column=1).value).encode())
		q.index=index			
		q.put()
		index+=1



app = get_wsgi_application()